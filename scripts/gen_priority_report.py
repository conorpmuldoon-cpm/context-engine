"""Generate a priority-ranked Excel report of all imported/blocked weblinks."""

import csv
import json
import glob
import re
from urllib.parse import urlparse, unquote
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

PROJECT_ROOT = Path(__file__).resolve().parent.parent

# --- Load data ---

def load_data():
    csv_path = PROJECT_ROOT / "cmuldoon emails input" / "filtered_weblinks.csv"
    state_path = PROJECT_ROOT / "outputs" / "link-importer-state.json"

    all_urls = []
    seen = set()
    with open(csv_path, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            url = row.get("url", "").strip().rstrip("#")
            if url and url not in seen:
                seen.add(url)
                all_urls.append(url)

    state = json.load(open(state_path))
    processed_urls = set(state.get("processed_urls", []))

    url_to_record = {}
    for pat in ["context-store/CTX-NEWS-*.json", "context-store/CTX-WEB-*.json"]:
        for fp in glob.glob(str(PROJECT_ROOT / pat)):
            try:
                rec = json.load(open(fp, encoding="utf-8"))
                src_url = rec.get("source_url", "")
                if src_url:
                    url_to_record[src_url] = {
                        "record_id": rec.get("record_id", ""),
                        "title": rec.get("title", ""),
                        "publication_date": rec.get("publication_date", ""),
                    }
            except Exception:
                pass

    return all_urls, processed_urls, url_to_record


# --- Priority scoring ---

HIGH_KEYWORDS = {
    "budget": 10, "tax": 10, "fiscal": 10, "spending": 10, "revenue": 9,
    "council": 10, "common-council": 10, "councilor": 9, "lawmaker": 9,
    "mayor": 10, "walsh": 10, "administration": 8,
    "police": 9, "spd": 9, "public-safety": 9, "crime": 7,
    "dpw": 9, "public-works": 9, "infrastructure": 9, "water": 8,
    "sewer": 8, "lead-pipe": 10, "lead": 7,
    "housing": 9, "affordable-housing": 10, "homelessness": 9,
    "homeless": 9, "shelter": 8,
    "development": 8, "economic-development": 9, "inner-harbor": 9,
    "micron": 10, "semiconductor": 9, "chips": 8, "workforce": 9,
    "union": 9, "contract": 8, "labor": 8,
    "zoning": 9, "planning": 8, "permit": 8, "variance": 8,
    "fire": 8, "firefighter": 8,
    "audit": 9, "oversight": 9, "accountability": 9, "transparency": 9,
    "equity": 9, "diversity": 8,
    "veto": 10, "override": 10, "resolution": 8, "ordinance": 9,
    "hire": 8, "appoint": 8, "retire": 8, "resign": 9, "chief": 8,
    "sanctuary": 9, "immigration": 8, "federal": 8, "trump": 7,
    "grant": 8, "funding": 8,
    "i-81": 9, "community-grid": 9, "nysdot": 8,
    "steam": 8, "school": 7, "education": 7,
    "climate": 8, "greenhouse": 8, "sustainability": 8, "solar": 7,
    "neighborhood": 7, "resident": 7, "constituent": 8,
    "blueprint-15": 9, "blueprint": 8, "innovation": 8,
    "speed-camera": 8, "bicycle": 7, "pedestrian": 7, "traffic": 7,
    "concrete-plant": 8, "environmental": 8,
}

LOW_KEYWORDS = {
    "sports": -5, "basketball-game": 2, "football": -5, "baseball": -5,
    "recipe": -8, "restaurant": -5, "food": -5, "dining": -5,
    "obituar": -8, "obits": -8,
    "weather": -5, "forecast": -5,
    "entertainment": -5, "movie": -5, "concert": -5,
    "deer": -3, "wildlife": -3,
    "parking": -3,
    "plan-a-visit": -8,
}

SECTION_SCORES = {
    "/politics/": 8, "/news/": 5, "/opinion/": 4, "/editorials/": 4,
    "/crime/": 5, "/business/": 4, "/micron/": 8,
    "/realestate-news/": 3, "/education/": 4,
}

SKIP_DOMAINS = {"obits.syracuse.com", "www.tasteofsyracuse.com", "www.outsyracuse.com"}


def extract_date_from_url(url):
    m = re.search(r"/(\d{4})/(\d{2})/", url)
    if m:
        return int(m.group(1)), int(m.group(2))
    m = re.search(r"/(\d{4})-(\d{2})", url)
    if m:
        return int(m.group(1)), int(m.group(2))
    return None, None


def score_url(url):
    path = unquote(urlparse(url).path).lower().replace("_", "-")
    slug = path.rstrip("/").split("/")[-1] if path.rstrip("/") else ""
    full_text = path + " " + slug

    score = 0

    for kw, pts in HIGH_KEYWORDS.items():
        if kw in full_text:
            score += pts

    for kw, pts in LOW_KEYWORDS.items():
        if kw in full_text:
            score += pts

    for section, pts in SECTION_SCORES.items():
        if section in path:
            score += pts

    year, month = extract_date_from_url(url)
    if year:
        if year >= 2026:
            score += 8
        elif year == 2025 and month and month >= 7:
            score += 6
        elif year == 2025:
            score += 4
        elif year == 2024:
            score += 2
        else:
            score -= 2

    domain = urlparse(url).netloc.lower()
    if "syracuse.com" in domain:
        score += 3
    elif domain in ("cnycentral.com", "www.cnycentral.com"):
        score += 2
    elif "waer.org" in domain or "wrvo.org" in domain:
        score += 2

    return max(score, 0)


def priority_label(score):
    if score >= 20:
        return "High"
    elif score >= 12:
        return "Medium"
    elif score >= 5:
        return "Low"
    return "Minimal"


def slug_from_url(url):
    path = unquote(urlparse(url).path).rstrip("/")
    slug = path.split("/")[-1] if path else ""
    slug = slug.replace("-", " ").replace(".html", "").replace(".htm", "").strip()
    return slug.title()[:80] if slug else ""


# --- Excel generation ---

STATUS_FILLS = {
    "Imported": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "Blocked (paywall)": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "Filtered (not relevant)": PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid"),
    "Skipped (irrelevant domain)": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
    "Not processed": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
}

PRIORITY_FILLS = {
    "High": PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid"),
    "Medium": PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid"),
    "Low": PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid"),
    "Minimal": PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid"),
}

PRIORITY_FONTS = {
    "High": Font(bold=True, color="FFFFFF"),
    "Medium": Font(bold=True, color="000000"),
    "Low": Font(color="000000"),
    "Minimal": Font(color="808080"),
}


def main():
    all_urls, processed_urls, url_to_record = load_data()

    # Build rows
    rows = []
    for url in all_urls:
        domain = urlparse(url).netloc.lower()

        if domain in SKIP_DOMAINS:
            status = "Skipped (irrelevant domain)"
        elif url in url_to_record:
            status = "Imported"
        elif url in processed_urls:
            if "syracuse.com" in domain and "obits" not in domain:
                status = "Blocked (paywall)"
            else:
                status = "Filtered (not relevant)"
        else:
            status = "Not processed"

        score = score_url(url)
        rec = url_to_record.get(url, {})

        rows.append({
            "url": url,
            "domain": domain,
            "slug": slug_from_url(url),
            "status": status,
            "score": score,
            "priority": priority_label(score),
            "record_id": rec.get("record_id", ""),
            "title": rec.get("title", ""),
            "publication_date": rec.get("publication_date", ""),
        })

    # Sort: blocked first, then by score desc
    status_order = {
        "Blocked (paywall)": 0, "Not processed": 1,
        "Filtered (not relevant)": 2, "Imported": 3,
        "Skipped (irrelevant domain)": 4,
    }
    rows.sort(key=lambda r: (status_order.get(r["status"], 5), -r["score"]))

    # --- Excel workbook ---
    wb = Workbook()
    ws = wb.active
    ws.title = "Link Priority Ranking"

    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")

    headers = [
        "Rank", "Priority", "Score", "Status",
        "Article (from URL)", "Domain", "URL",
        "Record ID", "Title (if imported)", "Date",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for i, row in enumerate(rows, 2):
        ws.cell(row=i, column=1, value=i - 1)

        p_cell = ws.cell(row=i, column=2, value=row["priority"])
        p_cell.fill = PRIORITY_FILLS.get(row["priority"], PatternFill())
        p_cell.font = PRIORITY_FONTS.get(row["priority"], Font())
        p_cell.alignment = Alignment(horizontal="center")

        ws.cell(row=i, column=3, value=row["score"])

        s_cell = ws.cell(row=i, column=4, value=row["status"])
        if row["status"] in STATUS_FILLS:
            s_cell.fill = STATUS_FILLS[row["status"]]

        ws.cell(row=i, column=5, value=row["slug"])
        ws.cell(row=i, column=6, value=row["domain"])
        ws.cell(row=i, column=7, value=row["url"])
        ws.cell(row=i, column=8, value=row["record_id"])
        ws.cell(row=i, column=9, value=row["title"])
        ws.cell(row=i, column=10, value=row["publication_date"])

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 7
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 55
    ws.column_dimensions["F"].width = 22
    ws.column_dimensions["G"].width = 70
    ws.column_dimensions["H"].width = 22
    ws.column_dimensions["I"].width = 55
    ws.column_dimensions["J"].width = 14
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:J{len(rows) + 1}"

    # --- Summary sheet ---
    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 14
    ws2.column_dimensions["D"].width = 14
    ws2.column_dimensions["E"].width = 14

    ws2.cell(row=1, column=1, value="Priority x Status Breakdown").font = Font(bold=True, size=13)

    statuses = ["Blocked (paywall)", "Filtered (not relevant)", "Imported", "Skipped"]
    priorities = ["High", "Medium", "Low", "Minimal"]

    ws2.cell(row=3, column=1, value="Priority").font = Font(bold=True)
    for j, s in enumerate(statuses):
        ws2.cell(row=3, column=j + 2, value=s.split(" (")[0]).font = Font(bold=True, size=10)

    for i, p in enumerate(priorities):
        cell = ws2.cell(row=4 + i, column=1, value=p)
        cell.fill = PRIORITY_FILLS[p]
        cell.font = PRIORITY_FONTS[p]
        for j, s_label in enumerate(statuses):
            count = sum(
                1 for r in rows
                if r["priority"] == p and (
                    r["status"] == s_label or
                    (s_label == "Skipped" and r["status"] == "Skipped (irrelevant domain)")
                )
            )
            ws2.cell(row=4 + i, column=j + 2, value=count if count else "")

    ws2.cell(row=9, column=1, value="Total").font = Font(bold=True)
    ws2.cell(row=9, column=2, value=sum(1 for r in rows if r["status"] == "Blocked (paywall)")).font = Font(bold=True)
    ws2.cell(row=9, column=3, value=sum(1 for r in rows if r["status"] == "Filtered (not relevant)")).font = Font(bold=True)
    ws2.cell(row=9, column=4, value=sum(1 for r in rows if r["status"] == "Imported")).font = Font(bold=True)
    ws2.cell(row=9, column=5, value=sum(1 for r in rows if "Skipped" in r["status"])).font = Font(bold=True)

    # High-priority blocked list
    ws2.cell(row=11, column=1, value="High-Priority Blocked Articles").font = Font(bold=True, size=12, color="CC0000")
    ws2.cell(row=12, column=1, value="These are the most valuable unimported articles:").font = Font(italic=True)

    ws2.cell(row=13, column=1, value="Article").font = Font(bold=True)
    ws2.cell(row=13, column=2, value="Priority").font = Font(bold=True)
    ws2.cell(row=13, column=3, value="Score").font = Font(bold=True)
    ws2.cell(row=13, column=4, value="URL").font = Font(bold=True)
    ws2.column_dimensions["D"].width = 80

    high_blocked = [r for r in rows if r["priority"] in ("High", "Medium") and r["status"] == "Blocked (paywall)"]
    for k, r in enumerate(high_blocked):
        row_num = 14 + k
        ws2.cell(row=row_num, column=1, value=r["slug"][:55])
        cell = ws2.cell(row=row_num, column=2, value=r["priority"])
        cell.fill = PRIORITY_FILLS[r["priority"]]
        cell.font = PRIORITY_FONTS[r["priority"]]
        ws2.cell(row=row_num, column=3, value=r["score"])
        ws2.cell(row=row_num, column=4, value=r["url"])

    # Save
    output_path = PROJECT_ROOT / "outputs" / "link-import-priority.xlsx"
    wb.save(str(output_path))
    print(f"Saved: {output_path}")
    print()

    print("Priority breakdown:")
    for p in priorities:
        total = sum(1 for r in rows if r["priority"] == p)
        blocked = sum(1 for r in rows if r["priority"] == p and r["status"] == "Blocked (paywall)")
        imported = sum(1 for r in rows if r["priority"] == p and r["status"] == "Imported")
        filtered_ = sum(1 for r in rows if r["priority"] == p and r["status"] == "Filtered (not relevant)")
        print(f"  {p:8s}: {total:3d} total | {imported:3d} imported | {blocked:3d} blocked | {filtered_:3d} filtered")

    print(f"\nHigh-priority blocked (worth pursuing): {sum(1 for r in rows if r['priority'] == 'High' and r['status'] == 'Blocked (paywall)')}")
    print(f"Medium-priority blocked: {sum(1 for r in rows if r['priority'] == 'Medium' and r['status'] == 'Blocked (paywall)')}")


if __name__ == "__main__":
    main()
